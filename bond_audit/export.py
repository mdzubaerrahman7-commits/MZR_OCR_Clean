"""Write pipeline outputs: cleaned structured data (JSON) and the audited
Excel workbook.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .audit import RowAudit

STATUS_FILL = {
    "OK": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "MISMATCH": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "INCOMPLETE": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}

DATA_COLUMNS = [
    ("sl_no", "SL No"),
    ("in_bond_no_date", "In Bond No & Date"),
    ("import_cno_date", "Import C-No & Date"),
    ("office_code", "Office Code"),
    ("item_no", "Item No"),
    ("material_name", "Name of Raw Materials"),
    ("unit", "Unit"),
    ("opening_qty", "Opening Qty"),
    ("consumption_with_wastage", "Consumption (with wastage)"),
    ("consumption_asycuda", "Consumption in KG for ASYCUDA"),
    ("closing_qty", "Closing Qty"),
]


def write_cleaned_data(
    output_path: Path,
    source_file: str,
    header_fields: dict,
    rows: list[dict],
    audits: list[RowAudit],
) -> None:
    payload = {
        "source_file": source_file,
        "header_fields": header_fields,
        "rows": rows,
        "audit": [
            {
                "row_index": a.row_index,
                "material_name": a.material_name,
                "status": a.status,
                "opening_qty": a.opening_qty,
                "consumption_with_wastage": a.consumption_with_wastage,
                "closing_qty": a.closing_qty,
                "expected_closing_qty": a.expected_closing_qty,
                "variance": a.variance,
                "notes": a.notes,
            }
            for a in audits
        ],
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2))


def write_excel_report(
    output_path: Path,
    source_file: str,
    header_fields: dict,
    rows: list[dict],
    audits: list[RowAudit],
) -> None:
    wb = Workbook()

    _write_data_sheet(wb.active, rows, audits)
    wb.active.title = "Extracted Data"
    _write_audit_sheet(wb.create_sheet("Audit Summary"), source_file, header_fields, audits)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_data_sheet(ws, rows: list[dict], audits: list[RowAudit]) -> None:
    header = [label for _, label in DATA_COLUMNS] + ["Audit Status", "Expected Closing Qty", "Variance", "Notes"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row, audit in zip(rows, audits):
        values = [row.get(key, "") for key, _ in DATA_COLUMNS]
        values += [audit.status, audit.expected_closing_qty, audit.variance, audit.notes]
        ws.append(values)
        fill = STATUS_FILL.get(audit.status)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    for i, _ in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.freeze_panes = "A2"


def _write_audit_sheet(ws, source_file: str, header_fields: dict, audits: list[RowAudit]) -> None:
    ws.append(["Source file", source_file])
    ws.append([])
    ws.append(["Header field", "Extracted value"])
    for key, value in header_fields.items():
        ws.append([key, value])
    ws.append([])

    ok = sum(1 for a in audits if a.status == "OK")
    mismatch = sum(1 for a in audits if a.status == "MISMATCH")
    incomplete = sum(1 for a in audits if a.status == "INCOMPLETE")
    ws.append(["Total line items", len(audits)])
    ws.append(["Reconciled OK", ok])
    ws.append(["Mismatches", mismatch])
    ws.append(["Incomplete (needs manual review)", incomplete])
    ws.append([])

    ws.append(["Row", "Material", "Status", "Notes"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for a in audits:
        if a.status == "OK":
            continue
        ws.append([a.row_index, a.material_name, a.status, a.notes])
        fill = STATUS_FILL.get(a.status)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 50
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
